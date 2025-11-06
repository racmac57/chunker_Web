"""
Celery Task Queue System for Chunker_v2
Handles async file processing with race condition prevention and batching
"""

import hashlib
import os
import json
import logging
import time
import traceback
import subprocess
import shutil
from typing import Dict, Any, List, Optional
from datetime import datetime
from pathlib import Path

try:
    from deduplication import DeduplicationManager
except ImportError:
    DeduplicationManager = None  # type: ignore

logger = logging.getLogger(__name__)

# Optional Celery imports with fallback
try:
    from celery import Celery, chain, group
    from celery.exceptions import Retry, WorkerLostError
    from celery.signals import task_prerun, task_postrun, task_failure
    CELERY_AVAILABLE = True
except ImportError:
    CELERY_AVAILABLE = False
    logger.warning("Celery not available. Processing functions will work without async tasks.")

try:
    import redis
    REDIS_AVAILABLE = True
except ImportError:
    REDIS_AVAILABLE = False
    logger.warning("Redis not available. Task coordination disabled.")

try:
    from watchdog.events import FileSystemEvent
    WATCHDOG_AVAILABLE = True
except ImportError:
    WATCHDOG_AVAILABLE = False
    logger.warning("Watchdog not available. File monitoring disabled.")

# Celery configuration (only if available)
CELERY_CONFIG = {
    'broker_url': os.getenv('REDIS_URL', 'redis://localhost:6379/0'),
    'result_backend': os.getenv('REDIS_URL', 'redis://localhost:6379/0'),
    'task_serializer': 'json',
    'accept_content': ['json'],
    'result_serializer': 'json',
    'timezone': 'UTC',
    'enable_utc': True,
    'task_track_started': True,
    'task_time_limit': 300,  # 5 minutes
    'task_soft_time_limit': 240,  # 4 minutes
    'worker_prefetch_multiplier': 1,
    'task_acks_late': True,
    'worker_disable_rate_limits': True,
}

DEDUP_MANAGER_CACHE: Dict[str, "DeduplicationManager"] = {}


def _compute_content_hash(text: str) -> str:
    normalized = (text or "").strip().lower()
    normalized = " ".join(normalized.split())
    return hashlib.sha256(normalized.encode("utf-8")).hexdigest()


def _dedup_cache_key(config: Dict[str, Any]) -> str:
    dedup_config = config.get("deduplication", {}) or {}
    persist_dir = config.get("chroma_persist_dir", "./chroma_db")
    payload = {
        "persist": persist_dir,
        "dedup": dedup_config,
    }
    try:
        return json.dumps(payload, sort_keys=True)
    except TypeError:
        # Remove non-serializable entries for cache key
        sanitized = {k: v for k, v in dedup_config.items() if isinstance(v, (str, int, float, bool, list, dict))}
        payload["dedup"] = sanitized
        return json.dumps(payload, sort_keys=True)


def _get_dedup_manager(config: Dict[str, Any]):
    if DeduplicationManager is None:
        return None

    dedup_config = config.get("deduplication", {}) or {}
    if not dedup_config.get("enabled"):
        return None

    cache_key = _dedup_cache_key(config)
    manager = DEDUP_MANAGER_CACHE.get(cache_key)
    if manager:
        return manager

    persist_dir = config.get("chroma_persist_dir", "./chroma_db")
    try:
        manager = DeduplicationManager(
            persist_directory=persist_dir,
            config=dedup_config,
            preload=True,
        )
    except Exception as exc:
        logger.warning("Failed to initialize deduplication manager: %s", exc)
        return None

    DEDUP_MANAGER_CACHE[cache_key] = manager
    return manager
# Initialize Celery app (only if available)
if CELERY_AVAILABLE:
    app = Celery('chunker_v2')
    app.config_from_object(CELERY_CONFIG)
else:
    app = None

# Redis client for coordination (only if available)
if REDIS_AVAILABLE:
    try:
        redis_client = redis.Redis.from_url(CELERY_CONFIG['broker_url'])
    except Exception as e:
        logger.warning(f"Redis connection failed: {e}")
        redis_client = None
else:
    redis_client = None

# Create a no-op decorator for when Celery is not available
if not CELERY_AVAILABLE:
    def celery_task_decorator(*args, **kwargs):
        """No-op decorator when Celery is not available"""
        def decorator(func):
            return func
        return decorator
    
    # Make app.task work when app is None
    class MockApp:
        def task(self, *args, **kwargs):
            return celery_task_decorator(*args, **kwargs)
    
    app = MockApp()

class TaskCoordinator:
    """
    Coordinates file processing tasks to prevent race conditions and enable batching.
    """
    
    def __init__(self, config: Dict[str, Any]):
        """
        Initialize task coordinator.
        
        Args:
            config: Configuration dictionary
        """
        self.config = config
        self.debounce_window = config.get("debounce_window", 1.0)
        self.max_workers = config.get("max_workers", 4)
        self.batch_size = config.get("batch_size", 10)
        self.failed_dir = config.get("failed_dir", "03_archive/failed")
        
        # Create failed directory
        os.makedirs(self.failed_dir, exist_ok=True)
        
        logger.info("Initialized TaskCoordinator")
    
    def should_process_file(self, file_path: str) -> bool:
        """
        Check if file should be processed based on debouncing rules.
        
        Args:
            file_path: Path to the file
            
        Returns:
            True if should process, False otherwise
        """
        try:
            current_time = time.time()
            file_key = f"file_processing:{file_path}"
            
            # Check if file is already being processed
            if redis_client.exists(f"processing:{file_path}"):
                logger.debug(f"File already being processed: {file_path}")
                return False
            
            # Check debounce window
            last_processed = redis_client.get(f"last_processed:{file_path}")
            if last_processed:
                last_time = float(last_processed.decode())
                if current_time - last_time < self.debounce_window:
                    logger.debug(f"File within debounce window: {file_path}")
                    return False
            
            # Mark as processing
            redis_client.setex(f"processing:{file_path}", 300, current_time)  # 5 min expiry
            
            return True
            
        except Exception as e:
            logger.error(f"Error checking file processing status: {e}")
            return False
    
    def mark_file_completed(self, file_path: str) -> None:
        """
        Mark file as completed processing.
        
        Args:
            file_path: Path to the file
        """
        try:
            current_time = time.time()
            redis_client.setex(f"last_processed:{file_path}", 3600, current_time)  # 1 hour expiry
            redis_client.delete(f"processing:{file_path}")
            logger.debug(f"Marked file as completed: {file_path}")
        except Exception as e:
            logger.error(f"Error marking file completed: {e}")
    
    def mark_file_failed(self, file_path: str, error: str) -> None:
        """
        Mark file as failed processing.
        
        Args:
            file_path: Path to the file
            error: Error message
        """
        try:
            redis_client.delete(f"processing:{file_path}")
            redis_client.setex(f"failed:{file_path}", 3600, error)  # 1 hour expiry
            logger.error(f"Marked file as failed: {file_path} - {error}")
        except Exception as e:
            logger.error(f"Error marking file failed: {e}")

# Task routing and configuration with priority support (only if Celery available)
if CELERY_AVAILABLE:
    CELERY_CONFIG.update({
        'task_routes': {
            'celery_tasks.process_file_task': {'queue': 'file_processing'},
            'celery_tasks.batch_process_files_task': {'queue': 'batch_processing'},
            'celery_tasks.add_to_rag_task': {'queue': 'rag_processing'},
            'celery_tasks.evaluate_rag_task': {'queue': 'evaluation'},
            'celery_tasks.health_check_task': {'queue': 'monitoring'},
            'celery_tasks.process_priority_file_task': {'queue': 'priority_processing'},
        },
        'task_annotations': {
            'celery_tasks.process_file_task': {'rate_limit': '10/m'},
            'celery_tasks.batch_process_files_task': {'rate_limit': '5/m'},
            'celery_tasks.add_to_rag_task': {'rate_limit': '20/m'},
            'celery_tasks.evaluate_rag_task': {'rate_limit': '2/m'},
            'celery_tasks.process_priority_file_task': {'rate_limit': '50/m'},  # Higher rate for priority
        },
        'beat_schedule': {
            'health-check': {
                'task': 'celery_tasks.health_check_task',
                'schedule': 60.0,  # Every minute
            },
            'cleanup-old-tasks': {
                'task': 'celery_tasks.cleanup_old_tasks',
                'schedule': 3600.0,  # Every hour
            },
        },
    })
    
    # Re-initialize with updated config if needed
    if not isinstance(app, type('MockApp', (), {})):
        app = Celery('chunker_v2')
        app.config_from_object(CELERY_CONFIG)
    
    # Task signals for monitoring
    @task_prerun.connect
    def task_prerun_handler(sender=None, task_id=None, task=None, args=None, kwargs=None, **kwds):
        """Log task start."""
        logger.info(f"Task {task.name} started: {task_id}")

    @task_postrun.connect
    def task_postrun_handler(sender=None, task_id=None, task=None, args=None, kwargs=None, retval=None, state=None, **kwds):
        """Log task completion."""
        logger.info(f"Task {task.name} completed: {task_id} - State: {state}")

    @task_failure.connect
    def task_failure_handler(sender=None, task_id=None, exception=None, traceback=None, einfo=None, **kwds):
        """Log task failures."""
        logger.error(f"Task {sender.name} failed: {task_id} - {exception}")

@app.task(bind=True, max_retries=3, default_retry_delay=60)
def process_file_task(self, file_path: str, dest_path: Optional[str], 
                     event_type: str, config: Dict[str, Any]) -> Dict[str, Any]:
    """
    Celery task for processing a single file.
    
    Args:
        file_path: Path to the file
        dest_path: Destination path (for moved events)
        event_type: Type of event (moved, created)
        config: Configuration dictionary
        
    Returns:
        Processing result dictionary
    """
    coordinator = TaskCoordinator(config)
    
    try:
        # Check if file should be processed
        if not coordinator.should_process_file(file_path):
            return {"status": "skipped", "reason": "debounced_or_processing"}
        
        # Process the file
        result = process_file_with_rag(file_path, dest_path, event_type, config)
        
        # Mark as completed
        coordinator.mark_file_completed(file_path)
        
        return result
        
    except Exception as e:
        logger.error(f"Task failed for {file_path}: {e}")
        
        # Mark as failed
        coordinator.mark_file_failed(file_path, str(e))
        
        # Retry if possible
        if self.request.retries < self.max_retries:
            logger.info(f"Retrying task for {file_path} (attempt {self.request.retries + 1})")
            raise self.retry(countdown=60 * (self.request.retries + 1))
        
        return {"status": "failed", "error": str(e)}

@app.task(bind=True, max_retries=3, default_retry_delay=30, priority=9)
def process_priority_file_task(self, file_path: str, dest_path: Optional[str], 
                             event_type: str, config: Dict[str, Any]) -> Dict[str, Any]:
    """
    High-priority Celery task for urgent files (legal, police departments).
    
    Args:
        file_path: Path to the file
        dest_path: Destination path (for moved events)
        event_type: Type of event
        config: Configuration dictionary
        
    Returns:
        Processing results
    """
    try:
        logger.info(f"Processing HIGH PRIORITY file: {file_path}")
        
        # Same processing logic as regular task but with priority logging
        from watcher_splitter import process_file_enhanced
        
        file_path_obj = Path(file_path)
        success = process_file_enhanced(file_path_obj, config)
        
        if success:
            logger.info(f"✅ HIGH PRIORITY file processed successfully: {file_path}")
            return {"status": "success", "priority": "high", "message": "Priority file processed"}
        else:
            logger.error(f"❌ HIGH PRIORITY file processing failed: {file_path}")
            return {"status": "failed", "priority": "high", "message": "Priority file processing failed"}
            
    except Exception as e:
        logger.error(f"Priority task failed: {e}")
        if self.request.retries < self.max_retries:
            logger.info(f"Retrying priority task (attempt {self.request.retries + 1})")
            raise self.retry(countdown=30 * (self.request.retries + 1))
        return {"status": "failed", "priority": "high", "error": str(e)}

@app.task(bind=True, max_retries=2, default_retry_delay=30)
def add_to_rag_task(self, chunks: List[str], metadata: Dict[str, Any], 
                   config: Dict[str, Any]) -> Dict[str, Any]:
    """
    Add chunks to RAG system (ChromaDB or FAISS).
    
    Args:
        chunks: List of text chunks
        metadata: File metadata
        config: Configuration dictionary
        
    Returns:
        RAG operation results
    """
    try:
        if not config.get("rag_enabled", False):
            return {"status": "disabled", "message": "RAG not enabled"}
        
        # Try ChromaDB first, fallback to FAISS
        try:
            from rag_integration import ChromaRAG

            chroma_rag = ChromaRAG(
                persist_directory=config.get("chroma_persist_dir", "./chroma_db")
            )
            dedup_manager = _get_dedup_manager(config)

            results = []
            skipped_duplicates = 0

            for i, chunk in enumerate(chunks):
                chunk_metadata = {
                    "file_name": metadata["file_info"]["name"],
                    "file_type": metadata["file_info"]["extension"],
                    "chunk_index": i + 1,
                    "timestamp": datetime.now().isoformat(),
                    "keywords": extract_keywords(chunk),
                    "file_size": metadata["file_info"]["size"]
                }

                content_hash = None
                dedup_identifier = f"{chunk_metadata['file_name']}:{chunk_metadata['chunk_index']}"

                if dedup_manager:
                    try:
                        is_duplicate, content_hash, existing_ids = dedup_manager.is_duplicate(
                            chunk,
                            chunk_id=dedup_identifier,
                        )
                    except Exception as dedup_error:
                        logger.warning(
                            "Deduplication check failed for chunk %s: %s",
                            dedup_identifier,
                            dedup_error,
                        )
                        is_duplicate, existing_ids = False, []
                        content_hash = None
                    else:
                        if is_duplicate:
                            skipped_duplicates += 1
                            preview = existing_ids[:3]
                            if len(existing_ids) > 3:
                                preview.append("...")
                            logger.info(
                                "Deduplication skipped duplicate chunk %s (matches: %s)",
                                dedup_identifier,
                                preview if preview else "existing chunk",
                            )
                            continue

                if content_hash is None:
                    content_hash = _compute_content_hash(chunk)

                chunk_metadata["content_hash"] = content_hash

                chunk_id = chroma_rag.add_chunk(chunk, chunk_metadata)
                results.append(chunk_id)

                if dedup_manager:
                    dedup_manager.add_hash(content_hash, chunk_id)

            response = {"status": "success", "method": "chromadb", "chunk_ids": results}
            if dedup_manager:
                response["skipped_duplicates"] = skipped_duplicates
            return response
            
        except ImportError:
            # Fallback to FAISS
            from ollama_integration import initialize_ollama_rag
            rag_system = initialize_ollama_rag()
            
            chunk_metadatas = []
            for i, chunk in enumerate(chunks):
                chunk_metadata = {
                    "source_file": metadata["file_info"]["name"],
                    "file_type": metadata["file_info"]["extension"],
                    "chunk_index": i + 1,
                    "timestamp": datetime.now().isoformat(),
                    "keywords": extract_keywords(chunk),
                    "file_size": metadata["file_info"]["size"]
                }
                chunk_metadatas.append(chunk_metadata)
            
            rag_system.add_documents(chunks, chunk_metadatas)
            rag_system.save_index()
            
            return {"status": "success", "method": "faiss", "chunks_added": len(chunks)}
        
    except Exception as e:
        logger.error(f"RAG task failed: {e}")
        if self.request.retries < self.max_retries:
            logger.info(f"Retrying RAG task (attempt {self.request.retries + 1})")
            raise self.retry(countdown=30 * (self.request.retries + 1))
        return {"status": "failed", "error": str(e)}

@app.task(bind=True, max_retries=1, default_retry_delay=120)
def evaluate_rag_task(self, query: str, answer: str, context: str,
                     config: Dict[str, Any]) -> Dict[str, Any]:
    """
    Evaluate RAG performance for a query-answer pair.
    
    Args:
        query: Input query
        answer: Generated answer
        context: Retrieved context
        config: Configuration dictionary
        
    Returns:
        Evaluation results
    """
    try:
        if not config.get("rag_enabled", False):
            return {"status": "disabled", "message": "RAG evaluation not enabled"}
        
        from comprehensive_eval import ComprehensiveRAGEvaluator
        
        evaluator = ComprehensiveRAGEvaluator(config)
        results = evaluator.evaluate_end_to_end(query, answer, context)
        
        return {"status": "success", "evaluation": results}
        
    except Exception as e:
        logger.error(f"RAG evaluation task failed: {e}")
        if self.request.retries < self.max_retries:
            logger.info(f"Retrying evaluation task (attempt {self.request.retries + 1})")
            raise self.retry(countdown=120)
        return {"status": "failed", "error": str(e)}

@app.task
def health_check_task() -> Dict[str, Any]:
    """
    Health check task for monitoring system status.
    
    Returns:
        Health status information
    """
    try:
        # Check Redis connection
        redis_status = "healthy"
        try:
            redis_client.ping()
        except Exception as e:
            redis_status = f"unhealthy: {e}"
        
        # Check Celery worker status
        inspect = app.control.inspect()
        active_workers = inspect.active()
        worker_count = len(active_workers) if active_workers else 0
        
        # Check system resources
        import psutil
        cpu_percent = psutil.cpu_percent()
        memory_percent = psutil.virtual_memory().percent
        
        return {
            "timestamp": datetime.now().isoformat(),
            "redis_status": redis_status,
            "worker_count": worker_count,
            "cpu_percent": cpu_percent,
            "memory_percent": memory_percent,
            "status": "healthy" if redis_status == "healthy" and worker_count > 0 else "degraded"
        }
        
    except Exception as e:
        logger.error(f"Health check failed: {e}")
        return {"status": "unhealthy", "error": str(e)}

@app.task
def cleanup_old_tasks() -> Dict[str, Any]:
    """
    Clean up old task results and temporary data.
    
    Returns:
        Cleanup results
    """
    try:
        # Clean up old Redis keys
        old_keys = redis_client.keys("file_processing:*")
        cleaned_count = 0
        
        for key in old_keys:
            try:
                redis_client.delete(key)
                cleaned_count += 1
            except Exception:
                continue
        
        return {
            "timestamp": datetime.now().isoformat(),
            "cleaned_keys": cleaned_count,
            "status": "success"
        }
        
    except Exception as e:
        logger.error(f"Cleanup task failed: {e}")
        return {"status": "failed", "error": str(e)}

def process_file_with_celery_chain(file_path: str, dest_path: Optional[str], 
                                 event_type: str, config: Dict[str, Any]) -> str:
    """
    Process file using Celery task chain.
    
    Args:
        file_path: Path to the file
        dest_path: Destination path (for moved events)
        event_type: Type of event
        config: Configuration dictionary
        
    Returns:
        Task ID for tracking
    """
    try:
        # Create task chain: process -> RAG -> evaluate (if enabled)
        if config.get("rag_enabled", False):
            # Chain: process_file -> add_to_rag -> evaluate_rag
            workflow = chain(
                process_file_task.s(file_path, dest_path, event_type, config),
                add_to_rag_task.s(config),
                evaluate_rag_task.s(config) if config.get("evaluation_enabled", False) else None
            )
        else:
            # Simple processing without RAG
            workflow = process_file_task.s(file_path, dest_path, event_type, config)
        
        # Execute the workflow
        result = workflow.apply_async()
        
        logger.info(f"Queued file processing workflow: {file_path} (task_id: {result.id})")
        return result.id
        
    except Exception as e:
        logger.error(f"Failed to queue file processing workflow: {e}")
        # Fallback to direct processing
        return process_file_task.delay(file_path, dest_path, event_type, config).id

@app.task
def batch_process_files_task(file_paths: List[str], config: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    Celery task for batch processing multiple files.
    
    Args:
        file_paths: List of file paths
        config: Configuration dictionary
        
    Returns:
        List of processing results
    """
    results = []
    
    for file_path in file_paths:
        try:
            result = process_file_task.delay(file_path, None, "batch", config)
            results.append({"file_path": file_path, "task_id": result.id})
        except Exception as e:
            logger.error(f"Failed to queue task for {file_path}: {e}")
            results.append({"file_path": file_path, "error": str(e)})
    
    return results

def process_file_with_rag(file_path: str, dest_path: Optional[str], 
                         event_type: str, config: Dict[str, Any]) -> Dict[str, Any]:
    """
    Process a file with RAG integration.
    
    Args:
        file_path: Path to the file
        dest_path: Destination path (for moved events)
        event_type: Type of event
        config: Configuration dictionary
        
    Returns:
        Processing result dictionary
    """
    try:
        file_path_obj = Path(file_path)
        
        # Check if file exists and is readable
        if not file_path_obj.exists():
            return {"status": "error", "message": "File does not exist"}
        
        if not os.access(file_path_obj, os.R_OK):
            return {"status": "error", "message": "File not readable"}
        
        # Extract metadata
        metadata = extract_comprehensive_metadata(file_path_obj)
        
        # Process file content
        content = extract_file_content(file_path_obj)
        
        if not content:
            return {"status": "error", "message": "No content extracted"}
        
        # Chunk the content
        chunks = chunk_content(content, config)
        
        if not chunks:
            return {"status": "error", "message": "No chunks created"}
        
        # Add to vector database
        vector_results = add_chunks_to_vector_db(chunks, metadata, config)
        
        # Create output files
        output_results = create_output_files(file_path_obj, chunks, config)
        
        # Handle file movement/archiving
        archive_result = archive_processed_file(file_path_obj, config)
        
        # Copy outputs back to source folder if needed
        if dest_path and event_type == "moved":
            copy_outputs_to_source(output_results, dest_path)
        
        return {
            "status": "success",
            "file_path": file_path,
            "chunks_created": len(chunks),
            "vector_results": vector_results,
            "output_files": output_results,
            "archive_result": archive_result,
            "metadata": metadata
        }
        
    except Exception as e:
        logger.error(f"Error processing file {file_path}: {e}")
        return {"status": "error", "message": str(e)}

def extract_comprehensive_metadata(file_path: Path) -> Dict[str, Any]:
    """
    Extract comprehensive metadata from a file.
    
    Args:
        file_path: Path to the file
        
    Returns:
        Metadata dictionary
    """
    try:
        stat = file_path.stat()
        
        metadata = {
            "file_info": {
                "name": file_path.name,
                "path": str(file_path),
                "size": stat.st_size,
                "created": datetime.fromtimestamp(stat.st_ctime).isoformat(),
                "modified": datetime.fromtimestamp(stat.st_mtime).isoformat(),
                "extension": file_path.suffix.lower(),
                "is_binary": is_binary_file(file_path)
            },
            "content_metadata": {
                "encoding": detect_encoding(file_path),
                "line_count": count_lines(file_path),
                "word_count": 0,  # Will be updated after content extraction
                "language": "unknown"  # Could be enhanced with language detection
            },
            "extracted_data": {
                "samples": [],
                "summaries": []
            },
            "tags": [],
            "keywords": [],
            "ai_context": {
                "summary": "",
                "category": "unknown",
                "complexity": "medium"
            }
        }
        
        return metadata
        
    except Exception as e:
        logger.error(f"Error extracting metadata from {file_path}: {e}")
        return {}

def extract_file_content(file_path: Path) -> str:
    """
    Extract content from various file types.
    
    Args:
        file_path: Path to the file
        
    Returns:
        Extracted content as string
    """
    try:
        extension = file_path.suffix.lower()
        
        if extension == ".xlsx":
            return extract_excel_content(file_path)
        elif extension == ".pdf":
            return extract_pdf_content(file_path)
        elif extension == ".py":
            return extract_python_content(file_path)
        elif extension == ".docx":
            return extract_docx_content(file_path)
        elif extension in [".txt", ".md", ".json", ".csv", ".sql", ".yaml", ".xml", ".log"]:
            return extract_text_content(file_path)
        else:
            logger.warning(f"Unsupported file type: {extension}")
            return ""
            
    except Exception as e:
        logger.error(f"Error extracting content from {file_path}: {e}")
        return ""

def extract_excel_content(file_path: Path) -> str:
    """Extract content from Excel files."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
        content = []
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            content.append(f"Sheet: {sheet_name}")
            
            # Extract data
            for row in sheet.iter_rows(values_only=True):
                if any(cell is not None for cell in row):
                    content.append(str(row))
            
            # Extract formulas
            formulas = []
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type == 'f' and cell.value:
                        formulas.append(f"{cell.coordinate}: {cell.value}")
            
            if formulas:
                content.append(f"Formulas: {formulas}")
        
        return "\n".join(content)
        
    except Exception as e:
        logger.error(f"Error extracting Excel content: {e}")
        return ""

def extract_pdf_content(file_path: Path) -> str:
    """Extract content from PDF files."""
    try:
        import PyPDF2
        content = []
        
        with open(file_path, 'rb') as f:
            reader = PyPDF2.PdfReader(f)
            
            for page_num, page in enumerate(reader.pages):
                text = page.extract_text()
                if text:
                    content.append(f"Page {page_num + 1}: {text}")
        
        return "\n".join(content)
        
    except Exception as e:
        logger.error(f"Error extracting PDF content: {e}")
        return ""

def extract_python_content(file_path: Path) -> str:
    """Extract structured content from Python files."""
    try:
        import ast
        
        with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
            content = f.read()
        
        # Parse AST
        tree = ast.parse(content)
        
        extracted = []
        
        # Extract functions
        functions = [node.name for node in ast.walk(tree) if isinstance(node, ast.FunctionDef)]
        if functions:
            extracted.append(f"Functions: {functions}")
        
        # Extract classes
        classes = [node.name for node in ast.walk(tree) if isinstance(node, ast.ClassDef)]
        if classes:
            extracted.append(f"Classes: {classes}")
        
        # Extract imports
        imports = []
        for node in ast.walk(tree):
            if isinstance(node, ast.Import):
                imports.extend([alias.name for alias in node.names])
            elif isinstance(node, ast.ImportFrom):
                imports.append(node.module)
        
        if imports:
            extracted.append(f"Imports: {imports}")
        
        # Add docstrings
        docstrings = []
        for node in ast.walk(tree):
            if isinstance(node, (ast.FunctionDef, ast.ClassDef, ast.Module)):
                if (ast.get_docstring(node)):
                    docstrings.append(f"{node.name}: {ast.get_docstring(node)}")
        
        if docstrings:
            extracted.append(f"Docstrings: {docstrings}")
        
        # Combine with original content
        extracted.append(f"Source Code:\n{content}")
        
        return "\n".join(extracted)
        
    except Exception as e:
        logger.error(f"Error extracting Python content: {e}")
        return ""

def extract_docx_content(file_path: Path) -> str:
    """Extract content from Word documents."""
    try:
        import docx
        doc = docx.Document(file_path)
        
        content = []
        for paragraph in doc.paragraphs:
            if paragraph.text.strip():
                content.append(paragraph.text)
        
        return "\n".join(content)
        
    except Exception as e:
        logger.error(f"Error extracting DOCX content: {e}")
        return ""

def extract_text_content(file_path: Path) -> str:
    """Extract content from text-based files."""
    try:
        with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
            return f.read()
    except Exception as e:
        logger.error(f"Error extracting text content: {e}")
        return ""

def chunk_content(content: str, config: Dict[str, Any]) -> List[str]:
    """
    Chunk content into manageable pieces.
    
    Args:
        content: Content to chunk
        config: Configuration dictionary
        
    Returns:
        List of content chunks
    """
    try:
        chunk_size = config.get("chunk_size", 1000)
        chunk_overlap = config.get("chunk_overlap", 200)
        
        # Simple chunking by character count with overlap
        chunks = []
        start = 0
        
        while start < len(content):
            end = min(start + chunk_size, len(content))
            chunk = content[start:end]
            
            if chunk.strip():
                chunks.append(chunk)
            
            start = end - chunk_overlap
        
        return chunks
        
    except Exception as e:
        logger.error(f"Error chunking content: {e}")
        return []

def add_chunks_to_vector_db(chunks: List[str], metadata: Dict[str, Any], 
                           config: Dict[str, Any]) -> Dict[str, Any]:
    """
    Add chunks to vector database.
    
    Args:
        chunks: List of content chunks
        metadata: File metadata
        config: Configuration dictionary
        
    Returns:
        Vector database operation results
    """
    try:
        if not config.get("rag_enabled", False):
            return {"status": "disabled"}
        
        # Try ChromaDB first, fallback to FAISS
        try:
            from rag_integration import ChromaRAG
            chroma_rag = ChromaRAG(persist_directory=config.get("chroma_persist_dir", "./chroma_db"))
            
            results = []
            for i, chunk in enumerate(chunks):
                chunk_metadata = {
                    "file_name": metadata["file_info"]["name"],
                    "file_type": metadata["file_info"]["extension"],
                    "chunk_index": i + 1,
                    "timestamp": datetime.now().isoformat(),
                    "keywords": extract_keywords(chunk),
                    "file_size": metadata["file_info"]["size"]
                }
                
                chunk_id = chroma_rag.add_chunk(chunk, chunk_metadata)
                results.append(chunk_id)
            
            return {"status": "success", "method": "chromadb", "chunk_ids": results}
            
        except ImportError:
            # Fallback to FAISS
            from ollama_integration import initialize_ollama_rag
            rag_system = initialize_ollama_rag()
            
            chunk_metadatas = []
            for i, chunk in enumerate(chunks):
                chunk_metadata = {
                    "source_file": metadata["file_info"]["name"],
                    "file_type": metadata["file_info"]["extension"],
                    "chunk_index": i + 1,
                    "timestamp": datetime.now().isoformat(),
                    "keywords": extract_keywords(chunk),
                    "file_size": metadata["file_info"]["size"]
                }
                chunk_metadatas.append(chunk_metadata)
            
            rag_system.add_documents(chunks, chunk_metadatas)
            rag_system.save_index()
            
            return {"status": "success", "method": "faiss", "chunks_added": len(chunks)}
        
    except Exception as e:
        logger.error(f"Error adding chunks to vector DB: {e}")
        return {"status": "error", "message": str(e)}

def create_output_files(file_path: Path, chunks: List[str], config: Dict[str, Any]) -> List[str]:
    """
    Create output files for chunks.
    
    Args:
        file_path: Original file path
        chunks: List of content chunks
        config: Configuration dictionary
        
    Returns:
        List of created output file paths
    """
    try:
        timestamp = datetime.now().strftime('%Y_%m_%d_%H_%M_%S')
        output_dir = Path(config['output_dir']) / f'{timestamp}_{file_path.stem}'
        output_dir.mkdir(parents=True, exist_ok=True)
        
        output_files = []
        
        for i, chunk in enumerate(chunks, 1):
            chunk_file = output_dir / f'{timestamp}_chunk_{i}.txt'
            with open(chunk_file, 'w', encoding='utf-8') as f:
                f.write(chunk)
            output_files.append(str(chunk_file))
        
        # Create transcript file
        transcript_file = output_dir / f'{timestamp}_transcript.md'
        with open(transcript_file, 'w', encoding='utf-8') as f:
            f.write(f"# Transcript: {file_path.name}\n\n")
            f.write(f"**Processing Time:** {timestamp}\n")
            f.write(f"**Chunks Created:** {len(chunks)}\n\n")
            f.write("## Content\n\n")
            f.write("\n\n---\n\n".join(chunks))
        
        output_files.append(str(transcript_file))
        
        return output_files
        
    except Exception as e:
        logger.error(f"Error creating output files: {e}")
        return []

def archive_processed_file(file_path: Path, config: Dict[str, Any]) -> str:
    """
    Enhanced archive with move operation, manifest validation, retry logic, and Git integration.
    
    Args:
        file_path: Path to the file
        config: Configuration dictionary
        
    Returns:
        Archive path
    """
    # Check if move_to_archive is enabled
    if not config.get('move_to_archive', False):
        logger.debug(f"Archive move disabled in config, skipping: {file_path.name}")
        return str(file_path)
    
    try:
        # Validate .origin.json manifest if it exists
        manifest_path = file_path.with_name(f"{file_path.name}.origin.json")
        manifest_data = None
        
        if manifest_path.exists():
            try:
                with open(manifest_path, 'r', encoding='utf-8') as f:
                    manifest_data = json.load(f)
                logger.info(f"Validated manifest for: {file_path.name}")
            except Exception as e:
                logger.warning(f"Failed to load manifest for {file_path.name}: {e}")
        else:
            logger.debug(f"No manifest found for: {file_path.name}")
        
        # Determine department from manifest or use default
        department = "admin"  # default
        if manifest_data and "department" in manifest_data:
            department = manifest_data["department"]
        
        # Create department-specific archive folder
        archive_dir = Path(config['archive_dir']) / department
        archive_dir.mkdir(parents=True, exist_ok=True)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        archive_path = archive_dir / f"{file_path.stem}_{timestamp}{file_path.suffix}"
        
        # Handle duplicate names
        counter = 1
        while archive_path.exists():
            archive_path = archive_dir / f"{file_path.stem}_{timestamp}_{counter}{file_path.suffix}"
            counter += 1
        
        # Retry MOVE operation up to 3 times
        move_success = False
        for attempt in range(3):
            try:
                logger.info(f"Attempting MOVE to archive (attempt {attempt + 1}): {file_path.name}")
                
                # Use shutil.move for better error handling
                shutil.move(str(file_path), str(archive_path))
                move_success = True
                logger.info(f"Successfully moved to archive: {archive_path.name}")
                break
                
            except Exception as e:
                logger.warning(f"MOVE attempt {attempt + 1} failed for {file_path.name}: {e}")
                if attempt < 2:  # Don't sleep on last attempt
                    time.sleep(1)
        
        # Fallback to COPY if all MOVE attempts failed
        if not move_success:
            logger.error(f"All MOVE attempts failed, falling back to COPY: {file_path.name}")
            try:
                shutil.copy2(str(file_path), str(archive_path))
                logger.warning(f"Used COPY fallback for archive: {archive_path.name}")
            except Exception as copy_error:
                logger.error(f"COPY fallback also failed: {copy_error}")
                return ""
        
        # Move manifest file if it exists
        if manifest_path.exists():
            try:
                manifest_archive = archive_path.with_name(f"{archive_path.name}.origin.json")
                shutil.move(str(manifest_path), str(manifest_archive))
                logger.debug(f"Moved manifest to archive: {manifest_archive.name}")
            except Exception as e:
                logger.warning(f"Failed to move manifest: {e}")
        
        # Git integration: Commit archived file
        if config.get('git_enabled', False):
            try:
                git_commit_archive(archive_path, file_path.name, timestamp)
            except Exception as e:
                logger.warning(f"Git commit failed: {e}")
        
        return str(archive_path)
        
    except Exception as e:
        logger.error(f"Error archiving file {file_path.name}: {e}")
        return ""

def git_commit_archive(archive_path: Path, original_filename: str, timestamp: str) -> None:
    """
    Commit archived file to Git with proper message and tagging.
    
    Args:
        archive_path: Path to archived file
        original_filename: Original filename
        timestamp: Timestamp string
    """
    try:
        # Get archive directory (assume repo root is archive_dir parent)
        repo_root = archive_path.parent.parent.parent  # Go up to C:/_chunker
        git_dir = repo_root / '.git'
        
        if not git_dir.exists():
            logger.debug("Git repository not found, skipping commit")
            return
        
        # Change to repo root and commit
        result = subprocess.run(
            ['git', 'add', str(archive_path.relative_to(repo_root))],
            cwd=str(repo_root),
            capture_output=True,
            text=True
        )
        
        if result.returncode != 0:
            logger.warning(f"Git add failed: {result.stderr}")
            return
        
        # Commit with descriptive message
        commit_msg = f"Archive {original_filename} - {timestamp}"
        result = subprocess.run(
            ['git', 'commit', '-m', commit_msg],
            cwd=str(repo_root),
            capture_output=True,
            text=True
        )
        
        if result.returncode != 0:
            logger.warning(f"Git commit failed: {result.stderr}")
        else:
            logger.info(f"Git commit successful: {commit_msg}")
        
    except Exception as e:
        logger.warning(f"Git operation failed: {e}")

def copy_outputs_to_source(output_files: List[str], dest_path: str) -> None:
    """
    Copy output files back to source folder.
    
    Args:
        output_files: List of output file paths
        dest_path: Destination path
    """
    try:
        dest_dir = Path(dest_path).parent
        dest_dir.mkdir(parents=True, exist_ok=True)
        
        for output_file in output_files:
            src_path = Path(output_file)
            dst_path = dest_dir / src_path.name
            shutil.copy2(src_path, dst_path)
            logger.info(f"Copied output to source: {dst_path}")
        
    except Exception as e:
        logger.error(f"Error copying outputs to source: {e}")

def extract_keywords(text: str, max_keywords: int = 5) -> List[str]:
    """
    Extract keywords from text.
    
    Args:
        text: Input text
        max_keywords: Maximum number of keywords
        
    Returns:
        List of keywords
    """
    try:
        import re
        from collections import Counter
        
        # Simple keyword extraction
        words = re.findall(r'\b\w+\b', text.lower())
        stop_words = {'the', 'a', 'an', 'and', 'or', 'but', 'in', 'on', 'at', 'to', 'for', 'of', 'with', 'by'}
        keywords = [word for word in words if word not in stop_words and len(word) > 2]
        
        counter = Counter(keywords)
        return [word for word, _ in counter.most_common(max_keywords)]
        
    except Exception as e:
        logger.error(f"Error extracting keywords: {e}")
        return []

def is_binary_file(file_path: Path) -> bool:
    """Check if file is binary."""
    try:
        with open(file_path, 'rb') as f:
            chunk = f.read(1024)
            return b'\0' in chunk
    except Exception:
        return False

def detect_encoding(file_path: Path) -> str:
    """Detect file encoding."""
    try:
        import chardet
        with open(file_path, 'rb') as f:
            raw_data = f.read(10000)
            result = chardet.detect(raw_data)
            return result.get('encoding', 'utf-8')
    except Exception:
        return 'utf-8'

def count_lines(file_path: Path) -> int:
    """Count lines in file."""
    try:
        with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
            return sum(1 for _ in f)
    except Exception:
        return 0

# Celery worker startup
if __name__ == '__main__':
    app.start()
