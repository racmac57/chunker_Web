"""
File Processors Module for Chunker_v2
Handles processing of different file types with improved error handling
"""

import logging
import ast
import json
import re
from typing import List, Dict, Any, Optional
from pathlib import Path

# File processing imports with error handling
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import PyPDF2
    PYPDF2_AVAILABLE = True
except ImportError:
    PYPDF2_AVAILABLE = False

try:
    import docx
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False

try:
    import yaml
    YAML_AVAILABLE = True
except ImportError:
    YAML_AVAILABLE = False

logger = logging.getLogger(__name__)


def process_python_file(text: str) -> str:
    """
    Extract detailed code structure from Python text content with descriptive blocks
    
    Args:
        text: Python file content as string
        
    Returns:
        Formatted string with detailed code structure analysis
    """
    try:
        tree = ast.parse(text)
        result_parts = []
        
        # Process each top-level node
        for node in ast.walk(tree):
            if isinstance(node, ast.Module):
                continue  # Skip module wrapper
                
            elif isinstance(node, ast.Import):
                # Handle import statements
                imports = [alias.name for alias in node.names]
                result_parts.append(f"[IMPORT] {', '.join(imports)}")
                
            elif isinstance(node, ast.ImportFrom):
                # Handle from imports
                module = node.module or ""
                imports = [alias.name for alias in node.names]
                result_parts.append(f"[FROM IMPORT] from {module} import {', '.join(imports)}")
                
            elif isinstance(node, ast.ClassDef):
                # Handle class definitions
                class_name = node.name
                bases = [base.id if hasattr(base, 'id') else str(base) for base in node.bases]
                base_str = f"({', '.join(bases)})" if bases else ""
                
                result_parts.append(f"\n[CLASS] {class_name}{base_str}")
                
                # Extract class methods
                methods = []
                for item in node.body:
                    if isinstance(item, ast.FunctionDef):
                        methods.append(item.name)
                
                if methods:
                    result_parts.append(f"   Methods: {', '.join(methods)}")
                
                # Add docstring if present
                if (node.body and isinstance(node.body[0], ast.Expr) 
                    and isinstance(node.body[0].value, ast.Constant) 
                    and isinstance(node.body[0].value.value, str)):
                    docstring = node.body[0].value.value.strip()
                    if docstring:
                        result_parts.append(f"   [Description] {docstring[:100]}...")
                
            elif isinstance(node, ast.FunctionDef):
                # Handle function definitions
                func_name = node.name
                args = [arg.arg for arg in node.args.args]
                args_str = f"({', '.join(args)})" if args else "()"
                
                # Determine function type
                if func_name.startswith('__') and func_name.endswith('__'):
                    func_type = "[MAGIC METHOD]"
                elif func_name.startswith('_'):
                    func_type = "[PRIVATE METHOD]"
                elif func_name.isupper():
                    func_type = "[CONSTANT FUNCTION]"
                else:
                    func_type = "[FUNCTION]"
                
                result_parts.append(f"\n{func_type}: {func_name}{args_str}")
                
                # Add docstring if present
                if (node.body and isinstance(node.body[0], ast.Expr) 
                    and isinstance(node.body[0].value, ast.Constant) 
                    and isinstance(node.body[0].value.value, str)):
                    docstring = node.body[0].value.value.strip()
                    if docstring:
                        result_parts.append(f"   [Description] {docstring[:100]}...")
                
                # Add decorators if present
                if node.decorator_list:
                    decorators = []
                    for decorator in node.decorator_list:
                        if hasattr(decorator, 'id'):
                            decorators.append(decorator.id)
                        elif hasattr(decorator, 'attr'):
                            decorators.append(decorator.attr)
                    result_parts.append(f"   [Decorators] {', '.join(decorators)}")
                
            elif isinstance(node, ast.Assign):
                # Handle variable assignments
                targets = []
                for target in node.targets:
                    if isinstance(target, ast.Name):
                        targets.append(target.id)
                
                if targets:
                    # Check if it's a constant (uppercase)
                    if all(name.isupper() for name in targets):
                        result_parts.append(f"[CONSTANT]: {', '.join(targets)}")
                    else:
                        result_parts.append(f"[VARIABLE]: {', '.join(targets)}")
        
        # If no structured content found, return basic analysis
        if not result_parts:
            # Fallback to basic analysis
            functions = [node.name for node in ast.walk(tree) if isinstance(node, ast.FunctionDef)]
            classes = [node.name for node in ast.walk(tree) if isinstance(node, ast.ClassDef)]
            imports = []
            
            for node in ast.walk(tree):
                if isinstance(node, ast.Import):
                    imports.extend([alias.name for alias in node.names])
                elif isinstance(node, ast.ImportFrom):
                    imports.append(node.module or "")
            
            result_parts = [
                f"[IMPORTS]: {', '.join(imports)}" if imports else "[IMPORTS]: None",
                f"[CLASSES]: {', '.join(classes)}" if classes else "[CLASSES]: None", 
                f"[FUNCTION]S: {', '.join(functions)}" if functions else "[FUNCTION]S: None"
            ]
        
        return "\n".join(result_parts)
        
    except SyntaxError as e:
        logger.error(f"Syntax error in Python file: {e}")
        # Fallback to raw text with basic structure
        lines = text.split('\n')
        basic_structure = []
        
        for i, line in enumerate(lines, 1):
            stripped = line.strip()
            if stripped.startswith('def '):
                func_name = stripped.split('(')[0].replace('def ', '')
                basic_structure.append(f"[FUNCTION] (line {i}): {func_name}")
            elif stripped.startswith('class '):
                class_name = stripped.split('(')[0].replace('class ', '')
                basic_structure.append(f"🏗️ CLASS (line {i}): {class_name}")
            elif stripped.startswith('import ') or stripped.startswith('from '):
                basic_structure.append(f"📦 IMPORT (line {i}): {stripped}")
        
        return "\n".join(basic_structure) if basic_structure else text
        
    except Exception as e:
        logger.error(f"Failed to process Python file: {e}")
        return text


def process_excel_file(file_path: Path) -> str:
    """
    Extract data from Excel files with enhanced error handling
    
    Args:
        file_path: Path to Excel file
        
    Returns:
        Extracted content as string
    """
    if not OPENPYXL_AVAILABLE:
        logger.error("openpyxl not available for Excel processing")
        return ""
    
    try:
        # First, validate that it's actually an Excel file
        if not _is_valid_excel_file(file_path):
            logger.warning(f"File {file_path} is not a valid Excel file, attempting alternative processing")
            return _process_corrupted_excel_file(file_path)
        
        wb = openpyxl.load_workbook(file_path)
        content = []
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            content.append(f"Sheet: {sheet_name}")
            
            # Extract formulas, data types, etc.
            for row in sheet.iter_rows(values_only=True):
                if any(cell is not None for cell in row):  # Skip empty rows
                    content.append(str(row))
        
        return "\n".join(content)
        
    except Exception as e:
        logger.error(f"Failed to process Excel file {file_path.name}: {e}")
        # Try to extract any readable content
        return _process_corrupted_excel_file(file_path)


def _is_valid_excel_file(file_path: Path) -> bool:
    """
    Check if file is a valid Excel file by examining its header.
    
    Args:
        file_path: Path to the file
        
    Returns:
        True if valid Excel file, False otherwise
    """
    try:
        with open(file_path, 'rb') as f:
            header = f.read(8)
            # Excel files start with PK (ZIP signature) or have specific Excel signatures
            return header.startswith(b'PK') or header.startswith(b'\xd0\xcf\x11\xe0')
    except Exception:
        return False


def _process_corrupted_excel_file(file_path: Path) -> str:
    """
    Attempt to process a corrupted Excel file by reading it as text.
    
    Args:
        file_path: Path to the corrupted file
        
    Returns:
        Extracted content as string
    """
    try:
        logger.info(f"Attempting to process corrupted file as text: {file_path}")
        
        # Try different encodings
        encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
        
        for encoding in encodings:
            try:
                with open(file_path, 'r', encoding=encoding, errors='replace') as f:
                    content = f.read()
                    
                # Check if it looks like tab-separated or comma-separated data
                if '\t' in content[:100]:
                    logger.info(f"Detected TSV format in {file_path}")
                    return f"TSV Data (detected):\n{content[:1000]}..."  # Limit output
                elif ',' in content[:100]:
                    logger.info(f"Detected CSV format in {file_path}")
                    return f"CSV Data (detected):\n{content[:1000]}..."
                else:
                    return f"Text Data (detected):\n{content[:1000]}..."
                    
            except UnicodeDecodeError:
                continue
        
        # If all encodings fail, return error message
        return f"Could not decode file {file_path} with any supported encoding"
        
    except Exception as e:
        logger.error(f"Failed to process corrupted file {file_path}: {e}")
        return f"Error processing corrupted file: {e}"


def process_pdf_file(file_path: Path) -> str:
    """
    Extract text from PDF files
    
    Args:
        file_path: Path to PDF file
        
    Returns:
        Extracted text content
    """
    if not PYPDF2_AVAILABLE:
        logger.error("PyPDF2 not available for PDF processing")
        return ""
    
    try:
        content = []
        with open(file_path, 'rb') as f:
            reader = PyPDF2.PdfReader(f)
            for page in reader.pages:
                page_text = page.extract_text()
                if page_text.strip():
                    content.append(page_text)
        
        return "\n".join(content)
        
    except Exception as e:
        logger.error(f"Failed to process PDF file {file_path.name}: {e}")
        return ""


def process_docx_file(file_path: Path) -> str:
    """
    Extract text from DOCX files
    
    Args:
        file_path: Path to DOCX file
        
    Returns:
        Extracted text content
    """
    if not DOCX_AVAILABLE:
        logger.error("python-docx not available for DOCX processing")
        return ""
    
    try:
        doc = docx.Document(file_path)
        paragraphs = [para.text for para in doc.paragraphs if para.text.strip()]
        return "\n".join(paragraphs)
        
    except Exception as e:
        logger.error(f"Failed to process DOCX file {file_path.name}: {e}")
        return ""


def process_yaml_file(text: str) -> str:
    """
    Extract data from YAML text content
    
    Args:
        text: YAML file content as string
        
    Returns:
        JSON formatted content
    """
    if not YAML_AVAILABLE:
        logger.error("PyYAML not available for YAML processing")
        return text
    
    try:
        data = yaml.safe_load(text)
        return json.dumps(data, indent=2)
        
    except Exception as e:
        logger.error(f"Failed to process YAML content: {e}")
        return text


def process_xml_file(text: str) -> str:
    """
    Process XML text content
    
    Args:
        text: XML file content as string
        
    Returns:
        Processed XML content
    """
    try:
        # Basic XML processing - could be enhanced with xml.etree
        return text
        
    except Exception as e:
        logger.error(f"Failed to process XML content: {e}")
        return text


def process_log_file(text: str) -> str:
    """
    Process log file content with streaming for large files
    
    Args:
        text: Log file content as string
        
    Returns:
        Processed log content
    """
    try:
        # For large log files, we could implement streaming
        # For now, return the text as-is
        return text
        
    except Exception as e:
        logger.error(f"Failed to process log content: {e}")
        return text


def process_sql_file(text: str) -> str:
    """
    Process SQL file content
    
    Args:
        text: SQL file content as string
        
    Returns:
        Processed SQL content
    """
    try:
        # Basic SQL processing - could be enhanced with SQL parsing
        return text
        
    except Exception as e:
        logger.error(f"Failed to process SQL content: {e}")
        return text


def process_text_file(text: str, file_type: str) -> str:
    """
    Process basic text files (.txt, .md, .json, .csv)
    
    Args:
        text: File content as string
        file_type: File extension
        
    Returns:
        Processed content
    """
    try:
        if file_type == ".json":
            # Validate JSON and pretty print
            try:
                data = json.loads(text)
                return json.dumps(data, indent=2)
            except json.JSONDecodeError:
                logger.warning("Invalid JSON format, returning raw text")
                return text
        
        return text
        
    except Exception as e:
        logger.error(f"Failed to process text file: {e}")
        return text


def get_file_processor(file_type: str):
    """
    Get the appropriate processor function for a file type
    
    Args:
        file_type: File extension (e.g., '.py', '.xlsx')
        
    Returns:
        Processor function
    """
    processors = {
        ".py": process_python_file,
        ".xlsx": process_excel_file,
        ".xlsm": process_excel_file,
        ".pdf": process_pdf_file,
        ".docx": process_docx_file,
        ".yaml": process_yaml_file,
        ".xml": process_xml_file,
        ".log": process_log_file,
        ".sql": process_sql_file,
    }
    
    return processors.get(file_type, process_text_file)


def check_processor_dependencies() -> Dict[str, bool]:
    """
    Check which file processor dependencies are available
    
    Returns:
        Dictionary with dependency status
    """
    return {
        "openpyxl": OPENPYXL_AVAILABLE,
        "PyPDF2": PYPDF2_AVAILABLE,
        "python-docx": DOCX_AVAILABLE,
        "PyYAML": YAML_AVAILABLE,
    }


def redact_sensitive_data(text: str) -> str:
    """
    Redact sensitive information from text
    
    Args:
        text: Input text
        
    Returns:
        Text with sensitive data redacted
    """
    try:
        # Redact common sensitive patterns
        patterns = [
            (r'\b\d{3}-\d{2}-\d{4}\b', '[SSN]'),  # SSN
            (r'\b\d{4} \d{4} \d{4} \d{4}\b', '[CC]'),  # Credit card
            (r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b', '[EMAIL]'),  # Email
            (r'\b\d{3}-\d{3}-\d{4}\b', '[PHONE]'),  # Phone
        ]
        
        redacted_text = text
        for pattern, replacement in patterns:
            redacted_text = re.sub(pattern, replacement, redacted_text)
        
        return redacted_text
        
    except Exception as e:
        logger.error(f"Failed to redact sensitive data: {e}")
        return text


if __name__ == "__main__":
    # Test dependency checking
    deps = check_processor_dependencies()
    print("File Processor Dependencies:")
    for dep, status in deps.items():
        print(f"  {dep}: {'✓' if status else '✗'}")
    
    # Test Python processor
    test_python = """
def hello_world():
    print("Hello, World!")

class TestClass:
    def __init__(self):
        pass
"""
    result = process_python_file(test_python)
    print(f"\nPython processor test:\n{result}")
